import openpyxl


def getRowCount(file, sheetName):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheetName]
    return sheet.max_row


def getColCount(file, sheetName):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheetName]
    return sheet.max_column


def readData(file, sheetName, rownum, colnum):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheetName]
    return sheet.cell(row=rownum, column=colnum).value


def writeDate(file, sheetName, rownum, colnum, data):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheetName]
    sheet.cell(row=rownum, column=colnum).value = data
    wb.save(file)

